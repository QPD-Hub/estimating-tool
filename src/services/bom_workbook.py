from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd

REQUIRED_BOM_HEADERS = (
    "Part Number",
    "Revision",
    "Quantity",
    "Level",
)

REQUIRED_BOM_HEADER_ALIASES = {
    "Part Number": ("part number", "partnumber", "part no", "part number", "pn"),
    "Revision": ("revision", "rev"),
    "Quantity": ("quantity", "qty"),
    "Level": ("level", "bom level", "indent level", "lvl"),
}


class BomWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class BomIdentity:
    part_number: str
    revision: str

    @property
    def filename(self) -> str:
        return f"{self.part_number}_{self.revision}_BOM.xlsx"


def extract_bom_identity(filename: str, workbook_content: bytes) -> BomIdentity:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return _extract_xlsx_bom_identity(workbook_content)
    if suffix == ".xls":
        return _extract_xls_bom_identity(workbook_content)

    raise BomWorkbookError(f"Unsupported BOM workbook type: {suffix or filename}")


def _extract_xlsx_bom_identity(workbook_content: bytes) -> BomIdentity:
    try:
        workbook = load_workbook(
            filename=BytesIO(workbook_content),
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, OSError, ValueError, KeyError, EOFError) as exc:
        raise BomWorkbookError("Workbook could not be read.") from exc
    except Exception as exc:
        raise BomWorkbookError("Workbook could not be read.") from exc

    try:
        if "BOM" not in workbook.sheetnames:
            raise BomWorkbookError("Worksheet 'BOM' is required.")

        worksheet = workbook["BOM"]
        header_map, header_row_index = _find_header_map(worksheet.iter_rows(values_only=True))

        for row in worksheet.iter_rows(
            min_row=header_row_index + 1,
            values_only=True,
        ):
            level_value = row[header_map["Level"]]
            if not _is_level_zero(level_value):
                continue

            part_number = _stringify_cell(row[header_map["Part Number"]])
            revision = _stringify_cell(row[header_map["Revision"]])

            if not part_number:
                raise BomWorkbookError(
                    "Level 0 row is missing a value for 'Part Number'."
                )
            if not revision:
                raise BomWorkbookError("Level 0 row is missing a value for 'Revision'.")

            return BomIdentity(part_number=part_number, revision=revision)

        raise BomWorkbookError("No level 0 row exists on worksheet 'BOM'.")
    finally:
        workbook.close()


def _extract_xls_bom_identity(workbook_content: bytes) -> BomIdentity:
    try:
        workbook = xlrd.open_workbook(file_contents=workbook_content)
    except (xlrd.XLRDError, OSError, ValueError, EOFError) as exc:
        raise BomWorkbookError("Workbook could not be read.") from exc
    except Exception as exc:
        raise BomWorkbookError("Workbook could not be read.") from exc

    try:
        worksheet = workbook.sheet_by_name("BOM")
    except xlrd.biffh.XLRDError as exc:
        raise BomWorkbookError("Worksheet 'BOM' is required.") from exc

    rows = (_normalize_xls_row(worksheet.row_values(index)) for index in range(worksheet.nrows))
    header_map, header_row_index = _find_header_map(rows)

    for row_index in range(header_row_index, worksheet.nrows):
        row = _normalize_xls_row(worksheet.row_values(row_index))
        level_value = row[header_map["Level"]] if header_map["Level"] < len(row) else None
        if not _is_level_zero(level_value):
            continue

        part_number = _cell_from_row(row, header_map["Part Number"])
        revision = _cell_from_row(row, header_map["Revision"])

        if not part_number:
            raise BomWorkbookError("Level 0 row is missing a value for 'Part Number'.")
        if not revision:
            raise BomWorkbookError("Level 0 row is missing a value for 'Revision'.")

        return BomIdentity(part_number=part_number, revision=revision)

    raise BomWorkbookError("No level 0 row exists on worksheet 'BOM'.")


def _find_header_map(rows) -> tuple[dict[str, int], int]:
    best_missing_headers = list(REQUIRED_BOM_HEADERS)

    for index, row in enumerate(rows, start=1):
        header_map: dict[str, int] = {}
        for column_index, value in enumerate(row):
            normalized_header = _normalize_header_name(value)
            if not normalized_header:
                continue
            for required_header, aliases in REQUIRED_BOM_HEADER_ALIASES.items():
                if normalized_header in aliases and required_header not in header_map:
                    header_map[required_header] = column_index
                    break
        missing_headers = [
            header for header in REQUIRED_BOM_HEADERS if header not in header_map
        ]
        if not missing_headers:
            return header_map, index
        if len(missing_headers) < len(best_missing_headers):
            best_missing_headers = missing_headers

    raise BomWorkbookError(
        "Worksheet 'BOM' is missing required columns: "
        f"{', '.join(best_missing_headers)}"
    )


def _is_level_zero(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return value == 0
    if isinstance(value, str):
        return value.strip() == "0"
    return False


def _normalize_xls_row(row: list[object]) -> list[object]:
    normalized_row: list[object] = []

    for value in row:
        if isinstance(value, float) and value.is_integer():
            normalized_row.append(int(value))
            continue
        normalized_row.append(value)

    return normalized_row


def _cell_from_row(row: list[object], index: int) -> str:
    if index >= len(row):
        return ""
    return _stringify_cell(row[index])


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header_name(value: object) -> str:
    normalized = _stringify_cell(value).lower().replace("#", " number ")
    normalized = "".join(character if character.isalnum() else " " for character in normalized)
    return " ".join(normalized.split())
